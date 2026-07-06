#!/usr/bin/env python3
"""把商务回填的 golden_review 合并回 golden，生成 v2 + 差异报告（闭环）。

低风险：默认 dry-run（只出报告不写 golden）；只有 --write 才生成新 golden；
永不覆盖 base golden（默认 eval/golden_masanduo_v1.jsonl）。仅用标准库；xlsx 为可选。

用法：
    # 预演（只出报告）
    python scripts/import_golden_review.py
    # 正式写 v2
    python scripts/import_golden_review.py --write \
        --input business_review/golden_review.csv \
        --base-golden eval/golden_masanduo_v1.jsonl \
        --output eval/golden_masanduo_v2.jsonl

business_status 支持：confirmed/revise/delete/pending/add（含中文：确认/修改/删除/待定/新增）。
字段名与上一轮导出的表不完全一致时做兼容映射（见 docs/10-golden-import-guide.md）。
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

# ---- 字段别名（同时兼容：英文机器名 / 中文表头 / 旧字段）----
ALIASES: Dict[str, List[str]] = {
    "id": ["id", "编号"],
    "category": ["category", "类别"],
    "role": ["role", "角色"],
    "question": ["question", "问题"],
    "expected_answer_points": ["expected_answer_points", "标准答案要点"],
    "required_sources": ["required_sources", "依据来源"],
    "forbidden_terms": ["forbidden_terms", "禁用词"],
    "must_handoff": ["must_handoff", "是否转人工"],
    "ideal_reply": ["ideal_reply", "理想回答"],
    "needs_business_review": ["needs_business_review", "是否需业务确认"],
    "business_status": ["business_status", "审核状态", "status", "状态"],
    "business_comment": ["business_comment", "业务备注", "business_notes", "（旧）业务备注", "备注", "notes"],
    "revised_question": ["revised_question", "修正-问题"],
    "revised_expected_answer_points": ["revised_expected_answer_points", "修正-标准答案要点"],
    "revised_required_sources": ["revised_required_sources", "修正-依据来源"],
    "revised_forbidden_terms": ["revised_forbidden_terms", "修正-禁用词"],
    "revised_must_handoff": ["revised_must_handoff", "修正-是否转人工"],
    "revised_ideal_reply": ["revised_ideal_reply", "修正-理想回答",
                            "business_corrected_answer", "（旧）修正答案", "修正答案"],
    "reviewer": ["reviewer", "审核人"],
    "owner": ["owner", "负责人"],
    "reviewed_at": ["reviewed_at", "审核日期", "review_date", "（旧）审核日期"],
}

STATUS_MAP = {
    "confirmed": "confirmed", "确认": "confirmed", "通过": "confirmed", "正确": "confirmed",
    "revise": "revise", "修改": "revise", "修正": "revise",
    "delete": "delete", "删除": "delete", "不要": "delete", "作废": "delete",
    "pending": "pending", "待定": "pending", "待确认": "pending",
    "add": "add", "新增": "add",
}

_TRUE = {"true", "是", "y", "yes", "1", "要", "转人工"}
_FALSE = {"false", "否", "n", "no", "0", "不要", "不转人工"}

HIGH_RISK_CATEGORIES = {"红线/合规", "服务费/费用解释", "审核失败/资料缺失", "监管锁/锁机说明"}
INLINE_FORBIDDEN = ["贷款", "利息", "包过", "一定通过", "一定发货"]


def parse_list(v: Any) -> List[str]:
    """宽松解析成 list：JSON数组 / 多行 / 顿号/逗号/分号分隔 / 空。"""
    if v is None:
        return []
    if isinstance(v, list):
        return [str(x).strip() for x in v if str(x).strip()]
    s = str(v).strip()
    if not s:
        return []
    if s.startswith("[") and s.endswith("]"):
        try:
            arr = json.loads(s)
            if isinstance(arr, list):
                return [str(x).strip() for x in arr if str(x).strip()]
        except Exception:
            pass
    parts = re.split(r"[\n、，,;；]+", s)
    return [p.strip() for p in parts if p.strip()]


def parse_bool(v: Any, default: Optional[bool] = None) -> Optional[bool]:
    if v is None:
        return default
    s = str(v).strip().lower()
    if s == "":
        return default
    if s in _TRUE:
        return True
    if s in _FALSE:
        return False
    return default


def _get(row: Dict[str, Any], canonical: str) -> Optional[Any]:
    for alias in ALIASES.get(canonical, [canonical]):
        if alias in row and str(row.get(alias, "")).strip() != "":
            return row[alias]
    return None


# ---------- 读取 ----------

def load_base_golden(path: Path) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    with path.open("r", encoding="utf-8") as f:
        for ln in f:
            ln = ln.strip()
            if ln:
                rows.append(json.loads(ln))
    return rows


def read_review_csv(path: Path) -> List[Dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def read_review_xlsx(path: Path) -> Optional[List[Dict[str, Any]]]:
    try:
        from openpyxl import load_workbook
    except Exception:
        return None
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = [str(h).strip() if h is not None else "" for h in next(rows_iter)]
    except StopIteration:
        return []
    out: List[Dict[str, Any]] = []
    for r in rows_iter:
        rec = {header[i]: (r[i] if i < len(r) else None) for i in range(len(header))}
        out.append(rec)
    return out


def resolve_input(args: argparse.Namespace) -> Tuple[Optional[List[Dict]], str, List[str]]:
    """返回 (rows, source_desc, notes)。"""
    notes: List[str] = []
    if args.input:
        p = Path(args.input)
        if not p.exists():
            return None, str(p), [f"输入文件不存在: {p}"]
        if p.suffix.lower() == ".xlsx":
            rows = read_review_xlsx(p)
            if rows is None:
                return None, str(p), ["指定了 xlsx 但未安装 openpyxl，请改用 CSV 或 pip install openpyxl"]
            return rows, str(p), notes
        return read_review_csv(p), str(p), notes
    # 未指定：优先 xlsx（若可读），否则 csv
    xlsx = Path("business_review/golden_review.xlsx")
    csvp = Path("business_review/golden_review.csv")
    if xlsx.exists():
        rows = read_review_xlsx(xlsx)
        if rows is not None:
            return rows, str(xlsx), notes
        notes.append("检测到 xlsx 但无 openpyxl，回退读取 CSV。")
    if csvp.exists():
        return read_review_csv(csvp), str(csvp), notes
    return None, "(none)", ["未找到 business_review/golden_review.csv 或 .xlsx"]


# ---------- 合并 ----------

def merge(base: List[Dict], review: List[Dict], now: str) -> Dict[str, Any]:
    by_id: Dict[str, Dict] = {str(c.get("id")): dict(c) for c in base}
    order: List[str] = [str(c.get("id")) for c in base]

    errors: List[str] = []
    warnings: List[str] = []
    diffs: List[Dict[str, Any]] = []
    deleted: List[str] = []
    pending: List[str] = []
    counts = {"confirmed": 0, "revise": 0, "delete": 0, "pending": 0, "add": 0, "unreviewed": 0}
    delete_ids = set()
    added_ids: List[str] = []

    reviewed_ids = set()

    for i, row in enumerate(review, 1):
        rid = str(_get(row, "id") or "").strip()
        raw_status = str(_get(row, "business_status") or "").strip()
        if raw_status == "" and rid == "":
            continue  # 空行
        status = STATUS_MAP.get(raw_status.lower(), STATUS_MAP.get(raw_status))
        if raw_status == "":
            # 未填状态 → 视为未审核，跳过（不改）
            continue
        if status is None:
            errors.append(f"[行{i} id={rid or '?'}] 无法识别的 business_status: {raw_status}")
            continue

        reviewer = str(_get(row, "reviewer") or "").strip()
        owner = str(_get(row, "owner") or "").strip()
        reviewed_at = str(_get(row, "reviewed_at") or "").strip() or now
        comment = str(_get(row, "business_comment") or "").strip()

        if status == "add":
            _handle_add(row, rid, by_id, order, added_ids, reviewer, owner, reviewed_at,
                        comment, errors, warnings, counts, now)
            if rid:
                reviewed_ids.add(rid)
            continue

        if not rid:
            errors.append(f"[行{i}] {status} 缺少 id")
            continue
        if rid not in by_id:
            errors.append(f"[行{i} id={rid}] 该 id 不在 base golden 中（如为新增请用 business_status=add）")
            continue
        reviewed_ids.add(rid)
        case = by_id[rid]

        if status == "delete":
            delete_ids.add(rid)
            deleted.append(rid)
            counts["delete"] += 1
            continue

        # 公共回填
        if reviewer:
            case["reviewer"] = reviewer
        if owner:
            case["owner"] = owner
        case["reviewed_at"] = reviewed_at
        if comment:
            case["business_comment"] = comment

        if status == "confirmed":
            case["needs_business_review"] = False
            counts["confirmed"] += 1
        elif status == "pending":
            case["needs_business_review"] = True
            pending.append(rid)
            counts["pending"] += 1
        elif status == "revise":
            diff = _apply_revise(row, case)
            case["needs_business_review"] = False
            if diff:
                diffs.append({"id": rid, "changes": diff})
            counts["revise"] += 1

        _warn_case(case, warnings, prefix=f"id={rid}")

    # 未在 review 里出现的 base case：保持不变
    for cid in order:
        if cid not in reviewed_ids and cid not in delete_ids:
            counts["unreviewed"] += 1

    new_golden = [by_id[cid] for cid in order if cid not in delete_ids]

    # 新版内 id 唯一性兜底
    seen = set()
    for c in new_golden:
        cid = str(c.get("id"))
        if cid in seen:
            errors.append(f"新版 golden 出现重复 id: {cid}")
        seen.add(cid)

    return {
        "new_golden": new_golden,
        "errors": errors,
        "warnings": warnings,
        "diffs": diffs,
        "deleted": deleted,
        "pending": pending,
        "added_ids": added_ids,
        "counts": counts,
    }


def _apply_revise(row: Dict, case: Dict) -> List[str]:
    changes: List[str] = []
    field_map = [
        ("revised_question", "question", "str"),
        ("revised_expected_answer_points", "expected_answer_points", "list"),
        ("revised_required_sources", "required_sources", "list"),
        ("revised_forbidden_terms", "forbidden_terms", "list"),
        ("revised_ideal_reply", "ideal_reply", "str"),
        ("revised_must_handoff", "must_handoff", "bool"),
    ]
    for rev_key, tgt, typ in field_map:
        raw = _get(row, rev_key)
        if raw is None or str(raw).strip() == "":
            continue
        old = case.get(tgt)
        if typ == "list":
            new = parse_list(raw)
        elif typ == "bool":
            new = parse_bool(raw, default=case.get(tgt))
        else:
            new = str(raw).strip()
        if new != old:
            changes.append(f"{tgt}: {_short(old)} → {_short(new)}")
            case[tgt] = new
    return changes


def _handle_add(row, rid, by_id, order, added_ids, reviewer, owner, reviewed_at,
                comment, errors, warnings, counts, now) -> None:
    if not rid:
        errors.append("[add] 缺少 id（新增题必须给唯一 id）")
        return
    if rid in by_id:
        errors.append(f"[add id={rid}] id 已存在，请换一个（建议 {rid}_new）")
        return

    def pick(canon_rev: str, canon_base: str) -> Any:
        return _get(row, canon_rev) if _get(row, canon_rev) is not None else _get(row, canon_base)

    question = str(pick("revised_question", "question") or "").strip()
    category = str(_get(row, "category") or "").strip()
    role = str(_get(row, "role") or "").strip()
    points = parse_list(pick("revised_expected_answer_points", "expected_answer_points"))
    ideal = str(pick("revised_ideal_reply", "ideal_reply") or "").strip()

    missing = [n for n, v in [("question", question), ("category", category), ("role", role),
                              ("expected_answer_points", points), ("ideal_reply", ideal)] if not v]
    if missing:
        errors.append(f"[add id={rid}] 缺少必要字段: {', '.join(missing)}")
        return

    confirmed = STATUS_MAP.get(str(_get(row, "business_status") or "").lower()) == "add"
    nbr = not (reviewer and False)  # add 默认待复核
    case = {
        "id": rid, "category": category, "role": role, "question": question,
        "expected_answer_points": points,
        "required_sources": parse_list(pick("revised_required_sources", "required_sources")),
        "forbidden_terms": parse_list(pick("revised_forbidden_terms", "forbidden_terms")),
        "must_handoff": bool(parse_bool(pick("revised_must_handoff", "must_handoff"), default=False)),
        "ideal_reply": ideal,
        "needs_business_review": True,  # 新增默认待复核
    }
    if reviewer:
        case["reviewer"] = reviewer
    if owner:
        case["owner"] = owner
    case["reviewed_at"] = reviewed_at
    if comment:
        case["business_comment"] = comment
    by_id[rid] = case
    order.append(rid)
    added_ids.append(rid)
    counts["add"] += 1
    _warn_case(case, warnings, prefix=f"add id={rid}")


def _warn_case(case: Dict, warnings: List[str], prefix: str) -> None:
    if not case.get("forbidden_terms"):
        warnings.append(f"[{prefix}] forbidden_terms 为空")
    if not case.get("required_sources"):
        warnings.append(f"[{prefix}] required_sources 为空")
    if not case.get("reviewer"):
        warnings.append(f"[{prefix}] reviewer 为空")
    if not case.get("owner"):
        warnings.append(f"[{prefix}] owner 为空")
    if case.get("needs_business_review"):
        warnings.append(f"[{prefix}] needs_business_review 仍为 true")
    if case.get("category") in HIGH_RISK_CATEGORIES and "must_handoff" not in case:
        warnings.append(f"[{prefix}] 高风险类别但 must_handoff 未明确")
    ideal = str(case.get("ideal_reply", ""))
    for t in INLINE_FORBIDDEN:
        if t in ideal:
            warnings.append(f"[{prefix}] ideal_reply 疑似出现禁用词「{t}」(可能是否定语境，请人工确认)")


def _short(v: Any, n: int = 40) -> str:
    s = json.dumps(v, ensure_ascii=False) if isinstance(v, (list, dict, bool)) else str(v)
    return s if len(s) <= n else s[:n] + "…"


# ---------- 校验 ----------

def validate_p0(new_golden: List[Dict]) -> List[str]:
    errs: List[str] = []
    seen = set()
    for c in new_golden:
        cid = str(c.get("id") or "").strip()
        if not cid:
            errs.append("存在 id 为空的 case")
        elif cid in seen:
            errs.append(f"id 重复: {cid}")
        else:
            seen.add(cid)
        for f in ["question", "category", "role", "ideal_reply"]:
            if not str(c.get(f) or "").strip():
                errs.append(f"[id={cid}] 必填字段 {f} 为空")
        if not c.get("expected_answer_points"):
            errs.append(f"[id={cid}] expected_answer_points 为空")
    return errs


# ---------- 报告 ----------

def write_report(path: Path, ctx: Dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    c = ctx["counts"]
    L: List[str] = []
    L.append(f"# golden 导入报告 {ctx['ts']}")
    L.append("")
    L.append(f"- 输入文件: `{ctx['source']}`")
    L.append(f"- base golden: `{ctx['base']}`")
    L.append(f"- 输出文件: `{ctx['output']}`（{'已写入' if ctx['written'] else 'dry-run 未写入'}）")
    L.append(f"- 新版总题数: {len(ctx['new_golden'])}")
    L.append("")
    L.append("## 统计")
    L.append(f"- confirmed: {c['confirmed']}")
    L.append(f"- revised: {c['revise']}")
    L.append(f"- deleted: {c['delete']}")
    L.append(f"- pending: {c['pending']}")
    L.append(f"- added: {c['add']}")
    L.append(f"- 未审核(保持原样): {c['unreviewed']}")
    L.append(f"- errors: {len(ctx['errors'])}　warnings: {len(ctx['warnings'])}")
    if ctx["compat_notes"]:
        L.append("")
        L.append("## 兼容说明")
        for n in ctx["compat_notes"]:
            L.append(f"- {n}")
    L.append("")
    L.append("## 错误（P0，阻止写入）")
    L += ([f"- ✗ {e}" for e in ctx["errors"]] or ["无 ✅"])
    L.append("")
    L.append("## 警告（P1，不阻止）")
    L += ([f"- ! {w}" for w in ctx["warnings"]] or ["无 ✅"])
    L.append("")
    L.append("## revised diff")
    if ctx["diffs"]:
        for d in ctx["diffs"]:
            L.append(f"- **{d['id']}**")
            for ch in d["changes"]:
                L.append(f"  - {ch}")
    else:
        L.append("无")
    L.append("")
    L.append("## deleted cases")
    L += ([f"- {i}" for i in ctx["deleted"]] or ["无"])
    L.append("")
    L.append("## pending cases（仍需确认）")
    L += ([f"- {i}" for i in ctx["pending"]] or ["无"])
    L.append("")
    L.append("## added cases")
    L += ([f"- {i}" for i in ctx["added_ids"]] or ["无"])
    L.append("")
    L.append("## 下一步建议")
    L.append(_suggest(ctx))
    path.write_text("\n".join(L), encoding="utf-8")


def _suggest(ctx: Dict[str, Any]) -> str:
    tips: List[str] = []
    if ctx["errors"]:
        tips.append("- 有 P0 错误，**未写入 v2**。请修表后重跑（先 dry-run）。")
    else:
        if not ctx["written"]:
            tips.append("- 校验通过。加 `--write` 即可生成 v2 golden。")
        else:
            tips.append("- v2 已生成。下一步：`python eval/bench_masanduo.py --golden " + ctx["output"] + "` 跑分对比。")
    if ctx["pending"]:
        tips.append(f"- 仍有 {len(ctx['pending'])} 题 pending，需要业务/开会定。")
    if ctx["counts"]["unreviewed"]:
        tips.append(f"- 有 {ctx['counts']['unreviewed']} 题未审核（保持原样），建议后续补审。")
    return "\n".join(tips)


def main() -> None:
    ap = argparse.ArgumentParser(description="合并商务回填的 golden_review → v2 + 报告")
    ap.add_argument("--input", default="")
    ap.add_argument("--base-golden", default="eval/golden_masanduo_v1.jsonl")
    ap.add_argument("--output", default="eval/golden_masanduo_v2.jsonl")
    ap.add_argument("--report", default="")
    ap.add_argument("--write", action="store_true", help="真正写入 v2（默认 dry-run 只出报告）")
    args = ap.parse_args()

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    now_date = datetime.now().strftime("%Y-%m-%d")

    base_path = Path(args.base_golden)
    if not base_path.exists():
        print(f"[error] base golden 不存在: {base_path}", file=sys.stderr)
        sys.exit(1)
    base = load_base_golden(base_path)

    review, source, compat_notes = resolve_input(args)
    report_path = Path(args.report) if args.report else Path(
        f"business_review/golden_import_report_{ts}.md"
    )

    if review is None:
        # 仍出一份报告说明问题
        ctx = {
            "ts": ts, "source": source, "base": str(base_path), "output": args.output,
            "written": False, "new_golden": base, "errors": compat_notes, "warnings": [],
            "diffs": [], "deleted": [], "pending": [], "added_ids": [],
            "counts": {"confirmed": 0, "revise": 0, "delete": 0, "pending": 0, "add": 0, "unreviewed": len(base)},
            "compat_notes": compat_notes,
        }
        write_report(report_path, ctx)
        print(f"[error] 无法读取回填表。报告: {report_path}", file=sys.stderr)
        sys.exit(1)

    result = merge(base, review, now_date)
    p0 = validate_p0(result["new_golden"])
    errors = result["errors"] + p0

    written = False
    output_path = Path(args.output)
    if args.write and not errors:
        if output_path.resolve() == base_path.resolve():
            print("[refused] 拒绝覆盖 base golden。请用不同的 --output。", file=sys.stderr)
            sys.exit(2)
        with output_path.open("w", encoding="utf-8") as f:
            for c in result["new_golden"]:
                f.write(json.dumps(c, ensure_ascii=False) + "\n")
        written = True

    ctx = {
        "ts": ts, "source": source, "base": str(base_path), "output": args.output,
        "written": written, "new_golden": result["new_golden"],
        "errors": errors, "warnings": result["warnings"], "diffs": result["diffs"],
        "deleted": result["deleted"], "pending": result["pending"],
        "added_ids": result["added_ids"], "counts": result["counts"],
        "compat_notes": compat_notes,
    }
    write_report(report_path, ctx)

    c = result["counts"]
    print(f"输入: {source}")
    print(f"confirmed={c['confirmed']} revise={c['revise']} delete={c['delete']} "
          f"pending={c['pending']} add={c['add']} unreviewed={c['unreviewed']}")
    print(f"errors={len(errors)} warnings={len(result['warnings'])}")
    print(f"报告: {report_path}")
    if args.write and errors:
        print("[未写入] 存在 P0 错误，v2 未生成。修表后重试。")
    elif written:
        print(f"[已写入] {output_path}")
    else:
        print("[dry-run] 未写入 v2。加 --write 生成。")


if __name__ == "__main__":
    main()
