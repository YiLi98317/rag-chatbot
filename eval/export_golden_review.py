#!/usr/bin/env python3
"""把 golden 题库导出成商务可审核的表格（默认中文表头，CSV + 可选 xlsx）。

用途：让不懂代码/不懂英文的商务/客服/业务员在表格里逐条审核标准答案。
- 默认表头是**中文**；每列有中文批注；business_status 是中文下拉；另有"字段说明"页。
- 内部字段（英文 canonical）与 `scripts/import_golden_review.py` 对齐，导入脚本同时认中文和英文表头。

用法：
    python eval/export_golden_review.py                  # 中文表头（推荐给商务）
    python eval/export_golden_review.py --english-headers # 英文表头（机器/自测用）
    python eval/export_golden_review.py --include-legacy   # 附加旧字段列
"""

from __future__ import annotations

import argparse
import csv
import json
import sys
from pathlib import Path
from typing import Any, Dict, List

CORE = [
    "id", "category", "role", "question",
    "expected_answer_points", "required_sources", "forbidden_terms",
    "must_handoff", "ideal_reply", "needs_business_review",
]
REVIEW = ["business_status", "business_comment", "reviewer", "owner", "reviewed_at"]
REVISED = [
    "revised_question", "revised_expected_answer_points", "revised_required_sources",
    "revised_forbidden_terms", "revised_must_handoff", "revised_ideal_reply",
]
LEGACY = ["business_corrected_answer", "business_notes", "review_date"]

# canonical(英文机器名) → 中文列名。导入脚本 ALIASES 同时认这两种。
LABELS: Dict[str, str] = {
    "id": "编号", "category": "类别", "role": "角色", "question": "问题",
    "expected_answer_points": "标准答案要点", "required_sources": "依据来源",
    "forbidden_terms": "禁用词", "must_handoff": "是否转人工",
    "ideal_reply": "理想回答", "needs_business_review": "是否需业务确认",
    "business_status": "审核状态", "business_comment": "业务备注",
    "reviewer": "审核人", "owner": "负责人", "reviewed_at": "审核日期",
    "revised_question": "修正-问题", "revised_expected_answer_points": "修正-标准答案要点",
    "revised_required_sources": "修正-依据来源", "revised_forbidden_terms": "修正-禁用词",
    "revised_must_handoff": "修正-是否转人工", "revised_ideal_reply": "修正-理想回答",
    "business_corrected_answer": "（旧）修正答案", "business_notes": "（旧）业务备注",
    "review_date": "（旧）审核日期",
}

# 每列填写说明（用于批注 + 字段说明页）
NOTES: Dict[str, str] = {
    "id": "题目唯一编号，技术已填，勿改", "category": "题目类别，技术已填", "role": "面向角色，技术已填",
    "question": "真实问题，技术已填", "expected_answer_points": "标准答案应含的要点（技术草拟，勿直接改，改填“修正-标准答案要点”）",
    "required_sources": "该题应依据的知识（技术草拟）", "forbidden_terms": "该题绝不能出现的词（技术草拟）",
    "must_handoff": "是否必须转人工：是/否（技术草拟）", "ideal_reply": "理想回答（技术草拟，勿直接改，改填“修正-理想回答”）",
    "needs_business_review": "=是 表示技术拿不准，请优先审",
    "business_status": "【必填】下拉选：确认/修改/删除/待定/新增", "business_comment": "备注：为什么这样判/待确认点",
    "reviewer": "审核人姓名", "owner": "该条业务负责人", "reviewed_at": "审核日期，如 2026-07-04",
    "revised_question": "要改问题才填", "revised_expected_answer_points": "要改标准答案要点才填（多点换行或顿号分隔）",
    "revised_required_sources": "要改依据来源才填", "revised_forbidden_terms": "要改禁用词才填",
    "revised_must_handoff": "要改是否转人工才填：是/否", "revised_ideal_reply": "要改理想回答就填这里（最常用）",
    "business_corrected_answer": "旧字段，等同“修正-理想回答”，新表请用修正列", "business_notes": "旧字段，等同“业务备注”",
    "review_date": "旧字段，等同“审核日期”",
}

STATUS_VALUES_CN = ["确认", "修改", "删除", "待定", "新增"]
STATUS_VALUES_EN = ["confirmed", "revise", "delete", "pending", "add"]


def columns(include_legacy: bool) -> List[str]:
    """返回 canonical(英文) 列顺序。roundtrip/机器逻辑用它。"""
    return CORE + REVIEW + REVISED + (LEGACY if include_legacy else [])


def header_of(col: str, chinese: bool) -> str:
    return LABELS.get(col, col) if chinese else col


def _fmt_list(v: Any) -> str:
    if isinstance(v, list):
        return "\n".join(str(x) for x in v)
    return "" if v is None else str(v)


def load_golden(path: Path) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    with path.open("r", encoding="utf-8") as f:
        for ln in f:
            ln = ln.strip()
            if ln:
                rows.append(json.loads(ln))
    return rows


def to_row(case: Dict[str, Any], cols: List[str]) -> Dict[str, str]:
    base = {
        "id": str(case.get("id", "")),
        "category": str(case.get("category", "")),
        "role": str(case.get("role", "")),
        "question": str(case.get("question", "")),
        "expected_answer_points": _fmt_list(case.get("expected_answer_points")),
        "required_sources": _fmt_list(case.get("required_sources")),
        "forbidden_terms": _fmt_list(case.get("forbidden_terms")),
        "must_handoff": "是" if case.get("must_handoff") else "否",
        "ideal_reply": str(case.get("ideal_reply", "")),
        "needs_business_review": "是" if case.get("needs_business_review") else "否",
    }
    return {c: base.get(c, "") for c in cols}


def write_csv(rows: List[Dict[str, str]], cols: List[str], out: Path, chinese: bool) -> None:
    headers = [header_of(c, chinese) for c in cols]
    with out.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(headers)
        for r in rows:
            w.writerow([r.get(c, "") for c in cols])


def _help_rows(cols: List[str]) -> List[List[str]]:
    out = [["中文列名", "英文机器名", "怎么填"]]
    for c in cols:
        out.append([LABELS.get(c, c), c, NOTES.get(c, "")])
    out += [
        ["", "", ""],
        ["审核状态可填", "confirmed/revise/delete/pending/add", "中文：确认 / 修改 / 删除 / 待定 / 新增"],
        ["原则", "-", "涉费用/风控/红线宁可保守；拿不准就“待定”；不要为了让 AI 过测试而降低标准"],
    ]
    return out


def write_xlsx(rows: List[Dict[str, str]], cols: List[str], out: Path, chinese: bool) -> bool:
    try:
        from openpyxl import Workbook
        from openpyxl.comments import Comment
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.worksheet.datavalidation import DataValidation
    except Exception:
        return False
    wb = Workbook()
    ws = wb.active
    ws.title = "题库审核"
    ws.append([header_of(c, chinese) for c in cols])

    head_fill = PatternFill("solid", fgColor="D9E1F2")
    review_fill = PatternFill("solid", fgColor="FCE4D6")
    review_cols = set(REVIEW + REVISED + LEGACY)
    for idx, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=idx)
        cell.font = Font(bold=True)
        cell.fill = review_fill if col in review_cols else head_fill
        note = NOTES.get(col, "")
        if note:
            cell.comment = Comment(f"{note}\n(机器字段名: {col})", "system")
    for r in rows:
        ws.append([r.get(c, "") for c in cols])

    wrap_cols = {"expected_answer_points", "required_sources", "forbidden_terms",
                 "question", "ideal_reply", "revised_expected_answer_points",
                 "revised_ideal_reply", "revised_required_sources",
                 "revised_forbidden_terms", "business_comment", "business_corrected_answer"}
    widths = {"id": 12, "category": 16, "role": 8, "question": 34,
              "expected_answer_points": 40, "ideal_reply": 44, "business_status": 14,
              "revised_ideal_reply": 40, "revised_expected_answer_points": 34,
              "business_comment": 26}
    for idx, col in enumerate(cols, 1):
        letter = ws.cell(row=1, column=idx).column_letter
        ws.column_dimensions[letter].width = widths.get(col, 16)
        if col in wrap_cols:
            for cell in ws[letter][1:]:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"

    if "business_status" in cols and rows:
        bs_idx = cols.index("business_status") + 1
        letter = ws.cell(row=1, column=bs_idx).column_letter
        values = STATUS_VALUES_CN if chinese else STATUS_VALUES_EN
        dv = DataValidation(type="list", formula1='"%s"' % ",".join(values), allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{letter}2:{letter}{len(rows) + 1}")

    ws2 = wb.create_sheet("字段说明")
    for row in _help_rows(cols):
        ws2.append(row)
    for c in ws2[1]:
        c.font = Font(bold=True)
    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 30
    ws2.column_dimensions["C"].width = 64
    for r in ws2.iter_rows():
        for cell in r:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(out)
    return True


def main() -> None:
    ap = argparse.ArgumentParser(description="导出 golden 为商务可审核表（默认中文表头）")
    ap.add_argument("--golden", default="eval/golden_masanduo_v1.jsonl")
    ap.add_argument("--out-dir", default="business_review")
    ap.add_argument("--include-legacy", action="store_true", help="附加旧字段列")
    ap.add_argument("--english-headers", action="store_true", help="用英文机器表头（机器/自测用）")
    args = ap.parse_args()

    chinese = not args.english_headers
    golden_path = Path(args.golden)
    if not golden_path.exists():
        print(f"[error] golden 不存在: {golden_path}", file=sys.stderr)
        sys.exit(1)

    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    cols = columns(args.include_legacy)
    cases = load_golden(golden_path)
    rows = [to_row(c, cols) for c in cases]

    csv_path = out_dir / "golden_review.csv"
    write_csv(rows, cols, csv_path, chinese)
    print(f"[ok] CSV 已生成: {csv_path}（{len(rows)} 题，{len(cols)} 列，{'中文' if chinese else '英文'}表头）")

    xlsx_path = out_dir / "golden_review.xlsx"
    if write_xlsx(rows, cols, xlsx_path, chinese):
        print(f"[ok] Excel 已生成: {xlsx_path}（中文表头 + 表头批注 + 审核状态下拉 + 字段说明页）")
    else:
        print("[info] 未安装 openpyxl，跳过 Excel（CSV 已可用；如需 Excel: pip install openpyxl）")

    print("[提示] 审核状态填：确认/修改/删除/待定/新增；要改答案填“修正-理想回答”。")


if __name__ == "__main__":
    main()
